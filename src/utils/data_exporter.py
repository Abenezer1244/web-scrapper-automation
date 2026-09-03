"""Data exporter: CSV / Excel / JSON with CSV injection sanitization and R2 upload.

Lead CSV/Excel content (columns, phone normalization, name/address split,
sanitization) is owned by src/utils/lead_export.py so the in-app download and the
scheduled/R2 export produce the IDENTICAL dialer-ready file. This module owns only
file writing + R2 I/O. CSV, Excel AND JSON all go through the same canonical
builder, so every format carries identical keys/values for the same records.
The DNC/TCPA disclaimer is NOT written into the CSV/Excel (a disclaimer row breaks
dialer import) — it's surfaced in the delivery email body + download UI instead.
"""

import json
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import pandas as pd
import requests as _requests

from src.config import settings
from src.utils.lead_export import (
    LEAD_CSV_COLUMNS,
    _apply_visibility,
    build_lead_export_row,
    write_lead_csv,
)
from src.utils.lead_signals import auction_reference_date
from src.utils.logger import setup_logger

_logger = setup_logger("exporter")

# Amber header colour for Excel (matches BridgeLeads design system)
_AMBER_HEX = "F5A623"


def _r2_api_base() -> str:
    """Return the Cloudflare R2 API base URL for the configured account + bucket."""
    account_id = settings.R2_ACCOUNT_ID
    bucket = settings.R2_BUCKET_NAME
    return f"https://api.cloudflare.com/client/v4/accounts/{account_id}/r2/buckets/{bucket}"


def _r2_headers() -> dict[str, str]:
    """Return auth headers for the Cloudflare R2 API."""
    return {"Authorization": f"Bearer {settings.R2_API_TOKEN}"}


def _canonical_dataframe(
    records: list[Any], hidden_fields: set[str] | None = None,
    columns: list[str] | None = None,
) -> pd.DataFrame:
    """Build a DataFrame of canonical lead rows — the SAME columns + formatting
    (dialer split cols, normalized phones, sanitized values) as the CSV, so the
    Excel export matches the CSV exactly. No DNC footer; that lives in the email
    body + download UI (a disclaimer row breaks spreadsheet/dialer import).

    One `today` for the whole frame so a large export crossing UTC midnight can't
    give two rows different freshness_days/months_delinquent (Codex review).

    `hidden_fields` blanks the user-deselected hideable columns (header order is
    unchanged), keeping Excel byte-identical to the CSV for the same config.

    `columns` (from `resolve_lead_export_columns`) restricts the frame to a lean
    per-record-type subset, keeping Excel in lockstep with the CSV. None => full
    `LEAD_CSV_COLUMNS`. Selecting a subset of the full-width rows can't drift from
    the CSV because both project the SAME built rows.
    """
    # Both clocks from a SINGLE instant (see write_lead_csv).
    _now = datetime.now(UTC)
    today = _now.date()
    auction_today = auction_reference_date(_now)
    rows = [
        _apply_visibility(
            build_lead_export_row(r, today, auction_today=auction_today), hidden_fields
        )
        for r in records
    ]
    return pd.DataFrame(rows, columns=columns or LEAD_CSV_COLUMNS)


class DataExporter:
    """Export lead records to CSV / Excel / JSON and upload to Cloudflare R2."""

    def __init__(self, export_dir: str | None = None) -> None:
        self.export_dir = Path(export_dir) if export_dir else settings.EXPORTS_DIR
        self.export_dir.mkdir(parents=True, exist_ok=True)

    # ─── Local file export ────────────────────────────────────────────────────

    def to_csv(
        self, records: list[Any], filename: str = "export",
        hidden_fields: set[str] | None = None,
        columns: list[str] | None = None,
    ) -> Path:
        """Export records to the canonical dialer-ready CSV (shared builder).

        `columns` restricts the header to a lean per-record-type subset (None =
        full superset). Single-type callers resolve it via
        `resolve_lead_export_columns`; combined/batch callers omit it.
        """
        filepath = self._timestamped_path(filename, "csv")
        # newline="" so the csv writer doesn't emit blank lines between rows.
        with open(filepath, "w", encoding="utf-8", newline="") as f:
            write_lead_csv(records, f, hidden_fields=hidden_fields, columns=columns)
        _logger.info("CSV exported: %s (%d rows)", filepath.name, len(records))
        return filepath

    def to_excel(
        self, records: list[Any], filename: str = "export",
        hidden_fields: set[str] | None = None,
        columns: list[str] | None = None,
    ) -> Path:
        """Export records to an Excel file (canonical columns) with amber header.

        `columns` restricts the sheet to a lean per-record-type subset (None =
        full superset), keeping Excel in lockstep with the CSV.
        """
        filepath = self._timestamped_path(filename, "xlsx")
        df = _canonical_dataframe(records, hidden_fields, columns)

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Leads")
            ws = writer.sheets["Leads"]

            # Style header row: amber background, bold white text
            from openpyxl.styles import Alignment, Font, PatternFill
            header_fill = PatternFill(fill_type="solid", fgColor=_AMBER_HEX)
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Auto-fit column widths
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        _logger.info("Excel exported: %s (%d rows)", filepath.name, len(df))
        return filepath

    def to_json(
        self, records: list[Any], filename: str = "export",
        hidden_fields: set[str] | None = None,
        columns: list[str] | None = None,
    ) -> Path:
        """Export records to JSON (orient=records) — the CANONICAL lead schema.

        JSON now goes through the SAME builder as CSV/Excel
        (`build_lead_export_row`), so all three formats carry identical
        keys/values for the same records. This replaced an older path that dumped
        raw input dicts (leaking internal artifacts like `raw_html_hash` and the
        raw `enrichment_data` blob, with a schema that differed between the initial
        and enriched export passes).

        `hidden_fields` blanks the user-deselected hideable columns (and their
        dependents); `columns` (from `resolve_lead_export_columns`) applies the same
        lean per-record-type trim as the CSV. Values are already spreadsheet-safe —
        `build_lead_export_row` sanitizes each emitted field — so we do NOT sanitize
        again here (a second `sanitize_for_csv` pass would corrupt a formula-guarded
        value, e.g. turn ``'=cmd`` into ``''=cmd``).
        """
        filepath = self._timestamped_path(filename, "json")
        keys = columns or LEAD_CSV_COLUMNS
        # One consistent pair of "today"s for the whole file, from a SINGLE instant:
        # UTC for the tax signals, county-local for the auction countdown
        # (lead_signals.AUCTION_TZ).
        _now = datetime.now(UTC)
        today = _now.date()
        auction_today = auction_reference_date(_now)
        rows = []
        for rec in records:
            row = _apply_visibility(
                build_lead_export_row(rec, today, auction_today=auction_today), hidden_fields
            )
            # Project to the (possibly lean) column set, preserving canonical order.
            rows.append({k: row[k] for k in keys if k in row})
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(rows, f, ensure_ascii=False, indent=2, default=str)
        _logger.info("JSON exported: %s (%d rows)", filepath.name, len(rows))
        return filepath

    def export(
        self,
        records: list[dict[str, Any]],
        filename: str = "export",
        fmt: str | None = None,
        hidden_fields: set[str] | None = None,
        columns: list[str] | None = None,
    ) -> Path:
        """Export to the given format. Single entry point for all callers.

        `hidden_fields` (from `resolve_hidden_output_fields(config.fields)`) is
        forwarded to every format so the delivered file honors the user's output
        visibility selection identically across csv/json/excel.

        `columns` (from `resolve_lead_export_columns(record_type)`) selects the lean
        per-record-type column subset, forwarded to every format (csv/excel/json —
        all canonical now) so the delivered files stay identical. None = full
        superset (combined/batch callers omit it).
        """
        from src.config.constants import SUPPORTED_EXPORT_FORMATS
        fmt = (fmt or settings.EXPORT_FORMAT).lower()
        # Guard against drift: the accepted set is the shared
        # SUPPORTED_EXPORT_FORMATS constant (same one the schema validator +
        # worker use), so adding a format is a one-line constant change.
        if fmt not in SUPPORTED_EXPORT_FORMATS:
            raise ValueError(f"Unsupported export format: {fmt}")
        if fmt == "csv":
            return self.to_csv(records, filename, hidden_fields=hidden_fields, columns=columns)
        if fmt == "json":
            return self.to_json(records, filename, hidden_fields=hidden_fields, columns=columns)
        # excel | xlsx
        return self.to_excel(records, filename, hidden_fields=hidden_fields, columns=columns)

    # ─── R2 upload ────────────────────────────────────────────────────────────

    def upload_to_r2(self, local_path: Path, object_key: str) -> str:
        """Upload a local file to Cloudflare R2 and return the object key.

        Args:
            local_path: Path to the local file.
            object_key: S3-style key (e.g. 'exports/user_id/job_id/leads.csv').

        Returns:
            The object key stored in R2.

        Raises:
            ValueError: If object_key contains path traversal.
        """
        # Prevent path traversal attacks
        if ".." in object_key or object_key.startswith("/"):
            raise ValueError(f"Invalid object key: {object_key}")
        content_types = {
            ".csv": "text/csv",
            ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ".json": "application/json",
        }
        content_type = content_types.get(local_path.suffix, "application/octet-stream")

        url = f"{_r2_api_base()}/objects/{object_key}"
        headers = _r2_headers()
        headers["Content-Type"] = content_type

        with open(local_path, "rb") as f:
            resp = _requests.put(url, headers=headers, data=f, timeout=120)

        if resp.status_code not in (200, 201):
            raise RuntimeError(f"R2 upload failed ({resp.status_code}): {resp.text[:200]}")

        _logger.info("Uploaded to R2: %s", object_key)
        return object_key

    def get_download_url(self, object_key: str, expires_in: int = 3600) -> str:
        """Generate a temporary download URL for an R2 object.

        Strategy (in order):
        1. R2 public URL if configured
        2. S3-compatible presigned URL via boto3 (most reliable)
        3. Cloudflare R2 API presigned URL (requires ACCOUNT_ID)

        Args:
            object_key: The R2 object key.
            expires_in: URL expiry in seconds (default: 1hr for in-app, use 172800 for email).

        Returns:
            HTTPS download URL.
        """
        # Public-URL path requires an EXPLICIT opt-in (R2_ALLOW_PUBLIC_URLS).
        # Exports contain seller PII; a stray R2_PUBLIC_URL must not silently
        # hand out permanent unauthenticated links. Without the flag we fall
        # through to the presigned/streamed path below.
        if settings.R2_PUBLIC_URL and settings.R2_ALLOW_PUBLIC_URLS:
            return f"{settings.R2_PUBLIC_URL}/{object_key}"
        if settings.R2_PUBLIC_URL and not settings.R2_ALLOW_PUBLIC_URLS:
            _logger.warning(
                "R2_PUBLIC_URL is set but R2_ALLOW_PUBLIC_URLS is false — "
                "ignoring it and using presigned URLs (export PII safety)."
            )

        # S3-compatible presigned URL via boto3 against the R2 S3 endpoint.
        # This is the active production path on Railway (R2_ENDPOINT_URL +
        # R2_ACCESS_KEY_ID + R2_SECRET_ACCESS_KEY are the env vars set in
        # prod). Don't remove this branch as "legacy" without first
        # migrating prod onto either R2_PUBLIC_URL or R2_ACCOUNT_ID.
        if settings.R2_ENDPOINT_URL and settings.R2_ACCESS_KEY_ID and settings.R2_SECRET_ACCESS_KEY:
            try:
                import boto3
                from botocore.config import Config

                s3 = boto3.client(
                    "s3",
                    endpoint_url=settings.R2_ENDPOINT_URL,
                    aws_access_key_id=settings.R2_ACCESS_KEY_ID,
                    aws_secret_access_key=settings.R2_SECRET_ACCESS_KEY,
                    config=Config(signature_version="s3v4"),
                    region_name="auto",
                )
                presigned = s3.generate_presigned_url(
                    "get_object",
                    Params={"Bucket": settings.R2_BUCKET_NAME, "Key": object_key},
                    ExpiresIn=expires_in,
                )
                _logger.info("Generated S3 presigned URL for %s", object_key)
                return presigned
            except Exception as exc:
                _logger.warning("S3 presigned URL failed: %s", str(exc)[:80])

        # Cloudflare R2 native API presigned URL (used only when the
        # boto3 S3-compatible path above is not configured — currently
        # not the production path).
        if settings.R2_ACCOUNT_ID:
            url = f"{_r2_api_base()}/objects/{object_key}?presigned=true&expiresIn={expires_in}"
            try:
                resp = _requests.get(url, headers=_r2_headers(), timeout=30)
                if resp.status_code == 200:
                    data = resp.json()
                    presigned = data.get("result", {}).get("presignedUrl")
                    if presigned:
                        return presigned
            except Exception as exc:
                _logger.warning("R2 API presigned URL failed: %s", str(exc)[:80])

        _logger.error("No download URL method available for %s", object_key)
        raise RuntimeError("Export download is not configured. Contact support.")

    # ─── Helper ───────────────────────────────────────────────────────────────

    def _timestamped_path(self, base_name: str, extension: str) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return self.export_dir / f"{base_name}_{timestamp}.{extension}"
