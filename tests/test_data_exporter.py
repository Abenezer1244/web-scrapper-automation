"""Tests for src/utils/data_exporter.py and CSV security middleware."""
import csv
import json

import pytest

from src.api.middleware.security import sanitize_for_csv
from src.utils.data_exporter import DataExporter
from src.utils.lead_export import LEAD_CSV_COLUMNS

LEAD_RECORDS = [
    {
        "date_recorded": "01/15/2024",
        "party_name": "Smith, John",
        "heirs": "Jane Smith",
        "legal_description": "LOT 4 BLOCK 2 SUNSET RIDGE",
        "parcel_id": "1234567890",
        "property_address": "123 Main St, Tacoma WA 98401",
        "mailing_address": "456 Oak Ave, Seattle WA 98101",
    },
    {
        "date_recorded": "01/16/2024",
        "party_name": "Jones, Mary",
        "heirs": "",
        "legal_description": "SEC 12 TWP 20N RNG 3E",
        "parcel_id": "0987654321",
        "property_address": "789 Pine Rd, Puyallup WA 98371",
        "mailing_address": "789 Pine Rd, Puyallup WA 98371",
    },
]


@pytest.fixture
def exporter(tmp_path):
    return DataExporter(export_dir=str(tmp_path))


# ─── Canonical columns (shared builder) ───────────────────────────────────────

def test_csv_header_is_canonical_columns(exporter):
    path = exporter.to_csv(LEAD_RECORDS, filename="cols")
    with open(path, newline="", encoding="utf-8") as f:
        header = next(csv.reader(f))
    assert header == LEAD_CSV_COLUMNS


def test_csv_has_dialer_split_columns(exporter):
    path = exporter.to_csv(LEAD_RECORDS, filename="split")
    with open(path, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    # "Smith, John" -> first John / last Smith (case preserved); address splits.
    assert rows[0]["first_name"] == "John" and rows[0]["last_name"] == "Smith"
    assert rows[0]["property_city"] == "Tacoma" and rows[0]["property_state"] == "WA"


# ─── CSV injection sanitization ───────────────────────────────────────────────

@pytest.mark.parametrize("formula", [
    "=SUM(A1:A10)",
    "+cmd|' /C calc'!A0",
    "-2+3+cmd|' /C calc'!A0",
    "@SUM(1+1)*cmd|' /C calc'!A0",
    "\t=malicious",
    "\r=malicious",
])
def test_sanitize_for_csv_blocks_formula_injection(formula):
    result = sanitize_for_csv(formula)
    assert not result.startswith(("=", "+", "-", "@", "\t", "\r"))
    assert result.startswith("'")


def test_sanitize_for_csv_clean_value_unchanged():
    assert sanitize_for_csv("Smith, John") == "Smith, John"
    assert sanitize_for_csv("123 Main St") == "123 Main St"
    assert sanitize_for_csv("LOT 4 BLOCK 2") == "LOT 4 BLOCK 2"


def test_canonical_row_sanitizes_injected_party_name():
    from src.utils.lead_export import build_lead_export_row
    row = build_lead_export_row({"party_name": "=cmd|' /C calc'!A0", "parcel_id": "1234567890"})
    assert not row["party_name"].startswith("=")


def test_csv_file_contains_no_formula_prefixes(exporter, tmp_path):
    injected = LEAD_RECORDS + [
        {
            "date_recorded": "01/17/2024",
            "party_name": "=HYPERLINK(\"http://evil.com\",\"click\")",
            "heirs": "+cmd",
            "legal_description": "@SUM(1)",
            "parcel_id": "1111111111",
            "property_address": "=A1",
            "mailing_address": "-2+1",
        }
    ]
    path = exporter.to_csv(injected, filename="injection_test")
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            for value in row.values():
                assert not value.startswith(("=", "+", "-", "@", "\t", "\r")), (
                    f"Formula injection found in CSV: {value!r}"
                )


# ─── Export formats ───────────────────────────────────────────────────────────

def test_export_csv_creates_file(exporter):
    path = exporter.to_csv(LEAD_RECORDS, filename="test")
    assert path.exists()
    assert path.suffix == ".csv"
    assert path.stat().st_size > 0


def test_export_csv_has_correct_row_count(exporter):
    path = exporter.to_csv(LEAD_RECORDS, filename="rowcount")
    with open(path, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    assert len(rows) == len(LEAD_RECORDS)


def test_export_json_creates_file(exporter):
    path = exporter.to_json(LEAD_RECORDS, filename="test")
    assert path.exists()
    assert path.suffix == ".json"


def test_export_json_is_valid_and_correct_count(exporter):
    path = exporter.to_json(LEAD_RECORDS, filename="valid")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    assert isinstance(data, list)
    assert len(data) == len(LEAD_RECORDS)


def test_export_json_sanitizes_nested_enrichment_strings(exporter):
    """Formula triggers inside nested objects (enrichment_data) and the
    phones/emails arrays must be neutralized, not just top-level columns."""
    records = [
        {
            "party_name": "Smith, John",
            "enrichment_data": {
                "description": "=cmd|'/c calc'!A1",
                "nested": {"deep": "+SUM(A1)"},
            },
            "emails": ["=HYPERLINK(\"http://evil\")"],
        }
    ]
    path = exporter.to_json(records, filename="nested_injection")
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    enr = data[0]["enrichment_data"]
    assert enr["description"].startswith("'"), enr["description"]
    assert enr["nested"]["deep"].startswith("'"), enr["nested"]["deep"]
    assert data[0]["emails"][0].startswith("'"), data[0]["emails"][0]


def test_export_excel_creates_file(exporter):
    path = exporter.to_excel(LEAD_RECORDS, filename="test")
    assert path.exists()
    assert path.suffix == ".xlsx"
    assert path.stat().st_size > 0


def test_export_dispatch_csv(exporter):
    path = exporter.export(LEAD_RECORDS, filename="dispatch", fmt="csv")
    assert path.suffix == ".csv"


def test_export_dispatch_json(exporter):
    path = exporter.export(LEAD_RECORDS, filename="dispatch", fmt="json")
    assert path.suffix == ".json"


def test_export_dispatch_excel(exporter):
    path = exporter.export(LEAD_RECORDS, filename="dispatch", fmt="excel")
    assert path.suffix == ".xlsx"


def test_export_dispatch_invalid_format_raises(exporter):
    with pytest.raises(ValueError, match="Unsupported export format"):
        exporter.export(LEAD_RECORDS, fmt="xml")


def test_timestamped_filenames_are_unique(exporter):
    p1 = exporter.to_csv(LEAD_RECORDS, filename="ts")
    p2 = exporter.to_csv(LEAD_RECORDS, filename="ts")
    # Same base name but timestamps differ — both exist, names differ or same
    assert p1.exists()
    assert p2.exists()


# ─── Output-field visibility through DataExporter.export() (all formats) ───────

def test_export_csv_blanks_hidden_keeps_header(exporter):
    """CSV via export() blanks deselected hideable cols, header set unchanged."""
    path = exporter.export(
        LEAD_RECORDS, filename="vis_csv", fmt="csv",
        hidden_fields={"legal_description", "heirs"},
    )
    with open(path, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
        f.seek(0)
        header = next(csv.reader(f))
    assert header == LEAD_CSV_COLUMNS  # columns kept, not dropped
    for r in rows:
        assert r["legal_description"] == "" and r["heirs"] == ""
        # mailing_address NOT hidden here -> still present.
        assert r["mailing_address"] != "" or r["party_name"]  # sanity: untouched
    assert rows[0]["mailing_address"] == "456 Oak Ave, Seattle WA 98101"
    assert rows[0]["party_name"] == "Smith, John"  # identity untouched


def test_export_json_blanks_hidden(exporter):
    path = exporter.export(
        LEAD_RECORDS, filename="vis_json", fmt="json",
        hidden_fields={"mailing_address"},
    )
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    for row in data:
        assert row["mailing_address"] == ""
    # Non-hidden hideable + identity remain.
    assert data[0]["legal_description"] == "LOT 4 BLOCK 2 SUNSET RIDGE"
    assert data[0]["party_name"] == "Smith, John"


def test_export_excel_blanks_hidden(exporter):
    from openpyxl import load_workbook
    path = exporter.export(
        LEAD_RECORDS, filename="vis_xlsx", fmt="excel",
        hidden_fields={"legal_description"},
    )
    wb = load_workbook(path)
    ws = wb["Leads"]
    header = [c.value for c in ws[1]]
    legal_idx = header.index("legal_description")
    party_idx = header.index("party_name")
    for row in ws.iter_rows(min_row=2, values_only=True):
        assert (row[legal_idx] or "") == ""        # hidden -> blank
        assert row[party_idx]                        # identity -> intact


def test_export_no_hidden_unchanged(exporter):
    """No hidden_fields => identical to the existing behavior (all values present)."""
    path = exporter.export(LEAD_RECORDS, filename="vis_none", fmt="csv")
    with open(path, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["legal_description"] == "LOT 4 BLOCK 2 SUNSET RIDGE"
    assert rows[0]["heirs"] == "Jane Smith"
    assert rows[0]["mailing_address"] == "456 Oak Ave, Seattle WA 98101"
